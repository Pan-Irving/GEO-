import csv
import json
from io import StringIO
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


SUPPORTED_EXTENSIONS = {".md", ".txt", ".json", ".csv", ".xlsx", ".pdf", ".jpg", ".jpeg", ".png", ".webp"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}


class ParseError(ValueError):
    pass


def parse_material(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ParseError(f"暂不支持该文件格式：{suffix}")
    if suffix in {".md", ".txt"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return f"```json\n{json.dumps(data, ensure_ascii=False, indent=2)}\n```"
    if suffix == ".csv":
        return parse_csv(path)
    if suffix == ".xlsx":
        return parse_xlsx(path)
    if suffix == ".pdf":
        return parse_pdf(path)
    if suffix in IMAGE_EXTENSIONS:
        return parse_image_placeholder(path)
    raise ParseError(f"暂不支持该文件格式：{suffix}")


def parse_csv(path: Path) -> str:
    content = path.read_text(encoding="utf-8-sig", errors="ignore")
    reader = csv.reader(StringIO(content))
    rows = list(reader)
    return rows_to_markdown(rows, heading=path.name)


def parse_xlsx(path: Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sections: list[str] = []
    for sheet in workbook.worksheets:
        rows = [[cell if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
        sections.append(rows_to_markdown(rows, heading=sheet.title))
    return "\n\n".join(sections)


def parse_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    sections: list[str] = [f"## {path.name}", ""]
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        text = text.strip()
        if text:
            sections.append(f"### 第 {index} 页\n\n{text}")
    if len(sections) <= 2:
        sections.append("（PDF 未抽取到可读文本，可能是扫描件或图片型 PDF。请补充 OCR 后的文字资料。）")
    return "\n\n".join(sections)


def parse_image_placeholder(path: Path) -> str:
    return (
        f"## {path.name}\n\n"
        "（图片资料已纳入资料池，但当前版本不做本地 OCR。"
        "如图片包含证书、截图、市场地位声明或产品参数，请补充对应文字说明，"
        "否则 Agent 只能把它作为存在该图片资料的线索，不能当作可引用证据。）"
    )


def rows_to_markdown(rows: list[list[object]], heading: str) -> str:
    normalized = [[str(cell).strip() for cell in row] for row in rows if any(str(cell).strip() for cell in row)]
    if not normalized:
        return f"## {heading}\n\n（空表）"
    width = max(len(row) for row in normalized)
    padded = [row + [""] * (width - len(row)) for row in normalized]
    header = padded[0]
    body = padded[1:]
    lines = [f"## {heading}", "", "| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]
    lines.extend("| " + " | ".join(row) + " |" for row in body)
    return "\n".join(lines)
