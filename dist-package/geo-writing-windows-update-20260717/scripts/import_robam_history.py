#!/usr/bin/env python
"""Import Robam historical article library and self-media publication records.

This is a one-off migration script for the local GEO writing/publishing app.
It reads the four article-library workbooks as metadata truth, reads article
body files from the attachment folder, and rebuilds the publishing workbench
history from the self-media publication stats workbook.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import re
import shutil
import sqlite3
import sys
import uuid
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import UTC, datetime
from difflib import SequenceMatcher
from xml.etree import ElementTree
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PUBLISHING_BACKEND = ROOT / "publishing" / "backend"
sys.path.insert(0, str(PUBLISHING_BACKEND))

from app.security import hash_password  # noqa: E402


DEFAULT_ATTACHMENTS = Path("/Users/pan/Downloads/老板电器内容规划&发布_附件")
DEFAULT_PUBLICATION_STATS = Path("/Users/pan/Downloads/老板电器内容规划&发布_自媒体发布统计.xlsx")
DEFAULT_WEB_PUBLICATION_STATS = Path("/Users/pan/Downloads/老板电器内容规划&发布_网媒发布统计_发布记录.xlsx")
DEFAULT_LIBRARY_XLSX = [
    Path("/Users/pan/Downloads/老板电器内容规划&发布_补充文章库_新增文章库.xlsx"),
    Path("/Users/pan/Downloads/老板电器内容规划&发布_高端厨电文章库_高端厨电文章库.xlsx"),
    Path("/Users/pan/Downloads/老板电器内容规划&发布_名气文章库_名气文章库.xlsx"),
    Path("/Users/pan/Downloads/老板电器内容规划&发布_洗碗机文章库_洗碗机文章库.xlsx"),
]

PROJECTS_BY_CATEGORY = {
    "名气": ("名气GEO-c6f10f35", "名气GEO"),
    "洗碗机": ("老板电器洗碗机geo-60747392", "老板电器洗碗机geo"),
    "高端厨电": ("高端厨电GEO-b4879dad", "高端厨电GEO"),
}

AI_PLATFORM_MAP = {
    "deepseek": "DeepSeek",
    "DeepSeek": "DeepSeek",
    "豆包": "豆包",
    "千问": "千问",
    "元宝": "元宝",
    "Kimi": "Kimi",
    "kimi": "Kimi",
    "文心": "文心",
}

EMPLOYEE_USERNAME = {
    "王雨杰": "wangyujie",
    "孙艺榕": "sunyirong",
    "郑忆安": "zhengyian",
    "范雨涵": "fanyuhan",
    "刘英杰": "liuyingjie",
    "曾子渊": "zengziyuan",
    "黄经坤": "huangjingkun",
    "郑智权": "zhengzhiquan",
}


@dataclass
class ArticleMeta:
    article_id: str
    project_id: str
    project_name: str
    category: str
    keyword: str
    article_type: str
    title: str
    source_title: str
    source_file_value: str
    source_workbook: str
    source_sheet: str
    source_row: int
    creator: str = ""
    created_at: str = ""
    body_path: Path | None = None
    match_method: str = "missing"
    match_score: float = 0
    markdown: str = ""
    content_hash: str = ""
    placeholder: bool = False
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass
class PublicationRow:
    source_row: int
    employee_name: str
    published_at: str
    category: str
    ai_platforms: list[str]
    titles: list[str]
    media_name: str
    publish_url: str
    raw: dict[str, Any]


@dataclass
class WebPublicationRow:
    source_row: int
    employee_name: str
    published_at: str
    category: str
    ai_platforms: list[str]
    titles: list[str]
    media_name: str
    media_category: str
    media_kind: str
    publish_url: str
    actual_cost: float
    publisher: str
    settled: str
    raw: dict[str, Any]


def utc_now() -> str:
    return datetime.now(UTC).isoformat()


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_key(value: Any) -> str:
    text = normalize_text(value).strip().strip("《》")
    text = text.replace("i1 Pro", "i1Pro").replace("i1 pro", "i1Pro").replace("S2 Ultra", "S2Ultra")
    text = re.sub(r"^[0-9一二三四五六七八九十]+[._\-—、．]+", "", text)
    text = re.sub(
        r"^(支柱深度|支柱深度文章|支柱深度文|榜单推荐|榜单推荐文章|榜单推荐文|横评对比|横评对比文章|横评对比文|场景选购|场景选购文章|场景选购文|产品证据|产品证据文章|产品证据文|FAQ问答|FAQ问答短文|FQA)[文章文\-—：: ]*",
        "",
        text,
    )
    text = re.sub(r"[\s_：:，,？?！!、（）()\-—\"“”《》\[\]【】/\\]+", "", text)
    return text.lower()


def slug_hash(*parts: Any, size: int = 16) -> str:
    payload = "\u0001".join(normalize_text(part) for part in parts)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()[:size]


def read_sheet_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        for row_index in range(2, sheet.max_row + 1):
            row = {headers[column - 1]: sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1)}
            if not any(value is not None for value in row.values()):
                continue
            row["_source_workbook"] = path.name
            row["_source_sheet"] = sheet.title
            row["_source_row"] = row_index
            rows.append(row)
    return rows


def first_h1(path: Path) -> str:
    if path.suffix.lower() != ".md":
        return ""
    try:
        for line in path.read_text(encoding="utf-8", errors="ignore").splitlines()[:30]:
            if line.startswith("# "):
                return line[2:].strip()
    except OSError:
        return ""
    return ""


def collect_attachment_files(root: Path) -> list[Path]:
    return sorted(path for path in root.rglob("*") if path.is_file() and path.suffix.lower() in {".md", ".docx"})


def preferred_file(paths: list[Path], category: str, source_workbook: str) -> Path:
    if not paths:
        raise ValueError("paths cannot be empty")

    def score(path: Path) -> tuple[int, int, int]:
        directory = path.parent.name
        directory_score = 0
        if "补充" in source_workbook and directory == "补充文章库":
            directory_score = 3
        elif category and category in directory:
            directory_score = 2
        elif directory == "补充文章库":
            directory_score = 1
        suffix_score = 1 if path.suffix.lower() == ".md" else 0
        length_score = -len(path.name)
        return directory_score, suffix_score, length_score

    return max(paths, key=score)


def build_file_indexes(files: list[Path]) -> tuple[dict[str, list[Path]], dict[str, list[Path]]]:
    by_name: dict[str, list[Path]] = defaultdict(list)
    by_key: dict[str, list[Path]] = defaultdict(list)
    for path in files:
        by_name[path.name].append(path)
        by_key[normalize_key(path.name)].append(path)
        by_key[normalize_key(path.stem)].append(path)
        heading = first_h1(path)
        if heading:
            by_key[normalize_key(heading)].append(path)
    return by_name, by_key


def resolve_body_path(
    row: dict[str, Any],
    files: list[Path],
    by_name: dict[str, list[Path]],
    by_key: dict[str, list[Path]],
    low_threshold: float,
) -> tuple[Path | None, str, float]:
    category = normalize_text(row.get("品类"))
    source_workbook = normalize_text(row.get("_source_workbook"))
    file_value = normalize_text(row.get("文件"))
    titles = [normalize_text(row.get("文章标题")), normalize_text(row.get("标题"))]
    keys = [normalize_key(value) for value in [file_value, Path(file_value).stem if file_value else "", *titles] if value]

    if file_value in by_name:
        return preferred_file(by_name[file_value], category, source_workbook), "file_exact", 1

    contained = [path for path in files if path.name and path.name in file_value]
    if contained:
        return preferred_file(contained, category, source_workbook), "file_contains", 1

    normalized_matches: list[Path] = []
    for key in keys:
        normalized_matches.extend(by_key.get(key, []))
    if normalized_matches:
        return preferred_file(normalized_matches, category, source_workbook), "normalized_title_or_file", 1

    best_score = 0.0
    best_path: Path | None = None
    search_keys = [key for key in keys if key]
    for path in files:
        if category and category not in path.parent.name and path.parent.name != "补充文章库":
            continue
        candidate_keys = [normalize_key(path.stem), normalize_key(path.name), normalize_key(first_h1(path))]
        for source_key in search_keys:
            for candidate_key in candidate_keys:
                if not source_key or not candidate_key:
                    continue
                score = SequenceMatcher(None, source_key, candidate_key).ratio()
                if score > best_score:
                    best_score = score
                    best_path = path
    if best_path and best_score >= low_threshold:
        return best_path, "fuzzy", best_score
    return None, "placeholder", best_score


def markdown_from_docx(path: Path) -> str:
    try:
        from docx import Document
    except ImportError:
        return markdown_from_docx_xml(path)
    document = Document(path)
    lines: list[str] = []
    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        if not text:
            continue
        style_name = (paragraph.style.name if paragraph.style else "").lower()
        if "heading 1" in style_name or style_name in {"title", "标题 1"}:
            lines.append(f"# {text}")
        elif "heading 2" in style_name or style_name == "标题 2":
            lines.append(f"## {text}")
        else:
            lines.append(text)
    return "\n\n".join(lines).strip() + "\n"


def markdown_from_docx_xml(path: Path) -> str:
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        text = "".join(texts).strip()
        if text:
            paragraphs.append(text)
    markdown = "\n\n".join(paragraphs).strip()
    if markdown and not markdown.startswith("# "):
        markdown = f"# {path.stem}\n\n{markdown}"
    return markdown + "\n"


def read_article_markdown(article: ArticleMeta) -> str:
    if not article.body_path:
        return placeholder_markdown(article, "未找到可靠附件正文。")
    if article.body_path.suffix.lower() == ".md":
        return article.body_path.read_text(encoding="utf-8", errors="ignore").strip() + "\n"
    if article.body_path.suffix.lower() == ".docx":
        return markdown_from_docx(article.body_path)
    return placeholder_markdown(article, f"暂不支持的附件类型：{article.body_path.suffix}")


def placeholder_markdown(article: ArticleMeta, reason: str) -> str:
    title = article.title or article.source_title or "历史导入文章"
    return f"# {title}\n\n> {reason}\n\n关键词：{article.keyword}\n\n文章类型：{article.article_type}\n"


def build_articles(library_paths: list[Path], attachment_root: Path, low_threshold: float) -> list[ArticleMeta]:
    files = collect_attachment_files(attachment_root)
    by_name, by_key = build_file_indexes(files)
    articles: list[ArticleMeta] = []
    for library_path in library_paths:
        for row in read_sheet_rows(library_path):
            category = normalize_text(row.get("品类"))
            if category not in PROJECTS_BY_CATEGORY:
                continue
            keyword = normalize_text(row.get("优化关键词"))
            article_type = normalize_text(row.get("内容类型"))
            source_title = normalize_text(row.get("文章标题"))
            title = normalize_text(row.get("标题")) or source_title or normalize_text(row.get("文件"))
            if not keyword or not article_type or not title:
                continue
            project_id, project_name = PROJECTS_BY_CATEGORY[category]
            body_path, match_method, match_score = resolve_body_path(row, files, by_name, by_key, low_threshold)
            article_id = f"robam-{slug_hash(project_id, row.get('_source_workbook'), row.get('_source_sheet'), row.get('_source_row'), title, keyword, article_type)}"
            article = ArticleMeta(
                article_id=article_id,
                project_id=project_id,
                project_name=project_name,
                category=category,
                keyword=keyword,
                article_type=article_type,
                title=title,
                source_title=source_title,
                source_file_value=normalize_text(row.get("文件")),
                source_workbook=normalize_text(row.get("_source_workbook")),
                source_sheet=normalize_text(row.get("_source_sheet")),
                source_row=int(row.get("_source_row") or 0),
                creator=normalize_text(row.get("创建人")),
                created_at=normalize_text(row.get("创建时间")),
                body_path=body_path,
                match_method=match_method,
                match_score=match_score,
                placeholder=body_path is None,
                raw={key: normalize_text(value) for key, value in row.items() if not str(key).startswith("_")},
            )
            article.markdown = read_article_markdown(article)
            article.content_hash = hashlib.sha256(article.markdown.strip().encode("utf-8")).hexdigest()
            articles.append(article)
    return articles


def date_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return normalize_text(value)


def split_ai_platforms(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    output: list[str] = []
    for item in re.split(r"[,，、/\n]+", raw):
        key = item.strip()
        if not key:
            continue
        output.append(AI_PLATFORM_MAP.get(key, key))
    return list(dict.fromkeys(output))


def normalize_channel(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    parts = [part.strip() for part in re.split(r"[,，\n]+", text) if part.strip()]
    normalized = []
    for part in parts:
        if part == "今日头条":
            part = "头条号"
        normalized.append(part)
    return "、".join(dict.fromkeys(normalized))


def publication_media_name(row: dict[str, Any]) -> str:
    channel = normalize_channel(row.get("发文渠道"))
    if channel:
        return channel
    starred = normalize_text(row.get("发文渠道*"))
    if starred:
        return starred
    for key, value in row.items():
        if str(key).startswith("发布账号-") and normalize_text(value):
            return normalize_text(value)
    return "未标注渠道"


def read_publications(path: Path) -> list[PublicationRow]:
    rows = read_sheet_rows(path)
    publications: list[PublicationRow] = []
    for row in rows:
        link = normalize_text(row.get("发布链接"))
        if not link:
            continue
        titles = [
            normalize_text(row.get("发布标题")),
            normalize_text(row.get("洗碗机发布标题")),
            normalize_text(row.get("高端厨电发布标题")),
            normalize_text(row.get("名气发布标题")),
            normalize_text(row.get("增发文章标题")),
        ]
        publications.append(
            PublicationRow(
                source_row=int(row.get("_source_row") or 0),
                employee_name=normalize_text(row.get("提交人")) or normalize_text(row.get("账号归属")) or "系统管理员",
                published_at=date_text(row.get("发布日期")),
                category=normalize_text(row.get("优化品类")),
                ai_platforms=split_ai_platforms(row.get("目标AI渠道")),
                titles=[title for title in titles if title],
                media_name=publication_media_name(row),
                publish_url=link,
                raw={key: normalize_text(value) for key, value in row.items() if not str(key).startswith("_")},
            )
        )
    return publications


def web_media_category(value: Any) -> str:
    category = normalize_text(value)
    if category == "垂媒":
        return "垂直媒体"
    return "大众媒体"


def read_web_publications(path: Path) -> list[WebPublicationRow]:
    rows = read_sheet_rows(path)
    publications: list[WebPublicationRow] = []
    for row in rows:
        if normalize_text(row.get("是否出稿")) not in {"1", "是", "已出稿", "true", "True"}:
            continue
        link = normalize_text(row.get("发布链接"))
        if not link:
            continue
        titles = [
            normalize_text(row.get("发布标题")),
            normalize_text(row.get("洗碗机发布标题")),
            normalize_text(row.get("高端厨电发布标题")),
            normalize_text(row.get("名气发布标题")),
            normalize_text(row.get("洗碗机补充文章选择")),
            normalize_text(row.get("高端厨电补充文章选择")),
            normalize_text(row.get("名气补充文章选择")),
        ]
        raw_cost = row.get("发布价格")
        try:
            actual_cost = float(raw_cost or 0)
        except (TypeError, ValueError):
            actual_cost = 0
        media_category = web_media_category(row.get("发布渠道类别"))
        publications.append(
            WebPublicationRow(
                source_row=int(row.get("_source_row") or 0),
                employee_name=normalize_text(row.get("创建人")) or "系统管理员",
                published_at=date_text(row.get("创建时间")),
                category=normalize_text(row.get("优化品类")),
                ai_platforms=split_ai_platforms(row.get("所属AI信源")),
                titles=[title for title in titles if title],
                media_name=normalize_text(row.get("发布渠道")) or "未标注媒体",
                media_category=media_category,
                media_kind=media_category,
                publish_url=link,
                actual_cost=actual_cost,
                publisher=normalize_text(row.get("发稿方")),
                settled=normalize_text(row.get("是否已结算")),
                raw={key: normalize_text(value) for key, value in row.items() if not str(key).startswith("_")},
            )
        )
    return publications


def index_articles(articles: list[ArticleMeta]) -> dict[str, list[ArticleMeta]]:
    index: dict[str, list[ArticleMeta]] = defaultdict(list)
    for article in articles:
        keys = [
            article.title,
            article.source_title,
            article.source_file_value,
            Path(article.source_file_value).stem if article.source_file_value else "",
        ]
        if article.body_path:
            keys.extend([article.body_path.name, article.body_path.stem, first_h1(article.body_path)])
        for key in keys:
            normalized = normalize_key(key)
            if normalized:
                index[normalized].append(article)
    return index


def find_article_for_publication(
    publication: PublicationRow,
    articles: list[ArticleMeta],
    article_index: dict[str, list[ArticleMeta]],
    low_threshold: float,
) -> tuple[ArticleMeta | None, str, float]:
    for title in publication.titles:
        matches = article_index.get(normalize_key(title), [])
        if publication.category:
            category_matches = [article for article in matches if article.category == publication.category]
            if category_matches:
                return category_matches[0], "title_exact_category", 1
        if matches:
            return matches[0], "title_exact", 1

    best_score = 0.0
    best_article: ArticleMeta | None = None
    title_keys = [normalize_key(title) for title in publication.titles if title]
    for article in articles:
        if publication.category and article.category != publication.category:
            continue
        candidate_keys = [normalize_key(article.title), normalize_key(article.source_title)]
        if article.body_path:
            candidate_keys.extend([normalize_key(article.body_path.stem), normalize_key(first_h1(article.body_path))])
        for title_key in title_keys:
            for candidate_key in candidate_keys:
                if not title_key or not candidate_key:
                    continue
                score = SequenceMatcher(None, title_key, candidate_key).ratio()
                if score > best_score:
                    best_score = score
                    best_article = article
    if best_article and best_score >= low_threshold:
        return best_article, "title_fuzzy", best_score
    return None, "missing", best_score


def add_publication_placeholders(
    articles: list[ArticleMeta],
    publications: list[PublicationRow],
    low_threshold: float,
) -> tuple[list[ArticleMeta], list[tuple[PublicationRow, ArticleMeta, str, float]]]:
    article_index = index_articles(articles)
    linked: list[tuple[PublicationRow, ArticleMeta, str, float]] = []
    next_articles = list(articles)
    for publication in publications:
        article, method, score = find_article_for_publication(publication, next_articles, article_index, low_threshold)
        if article is None:
            category = publication.category if publication.category in PROJECTS_BY_CATEGORY else "名气"
            project_id, project_name = PROJECTS_BY_CATEGORY[category]
            title = publication.titles[0] if publication.titles else f"历史发布记录 {publication.source_row}"
            article = ArticleMeta(
                article_id=f"robam-pub-placeholder-{slug_hash(publication.source_row, title, publication.publish_url)}",
                project_id=project_id,
                project_name=project_name,
                category=category,
                keyword=f"历史导入-{category}",
                article_type="历史发布文",
                title=title,
                source_title=title,
                source_file_value="",
                source_workbook=DEFAULT_PUBLICATION_STATS.name,
                source_sheet="自媒体发布统计",
                source_row=publication.source_row,
                match_method="publication_placeholder",
                match_score=score,
                placeholder=True,
                raw={"publication_row": publication.raw},
            )
            article.markdown = read_article_markdown(article)
            article.content_hash = hashlib.sha256(article.markdown.strip().encode("utf-8")).hexdigest()
            next_articles.append(article)
            article_index = index_articles(next_articles)
            method = "publication_placeholder"
        linked.append((publication, article, method, score))
    return next_articles, linked


def project_file(project_id: str, data_root: Path) -> Path:
    return data_root / "projects" / project_id / "project.json"


def backup_paths(data_root: Path, publishing_db: Path, backup_root: Path) -> dict[str, str]:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    target = backup_root / f"robam-history-{timestamp}"
    target.mkdir(parents=True, exist_ok=False)
    copied: dict[str, str] = {}
    if publishing_db.exists():
        dest = target / "publishing.db"
        shutil.copy2(publishing_db, dest)
        copied[str(publishing_db)] = str(dest)
    for project_id in {project_id for project_id, _ in PROJECTS_BY_CATEGORY.values()}:
        src = project_file(project_id, data_root)
        if src.exists():
            dest = target / f"{project_id}.project.json"
            shutil.copy2(src, dest)
            copied[str(src)] = str(dest)
    return copied


def read_env_value(path: Path, key: str) -> str:
    if not path.exists():
        return ""
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        current_key, value = line.split("=", 1)
        if current_key.strip() == key:
            return value.strip().strip('"').strip("'")
    return ""


def backup_sql_database(database_url: str, backup_root: Path) -> dict[str, str]:
    from sqlalchemy import create_engine, text

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    target = backup_root / f"robam-history-sql-{timestamp}"
    target.mkdir(parents=True, exist_ok=False)
    backup_file = target / "publishing_tables.json"
    engine = create_engine(database_url, future=True, pool_pre_ping=True)
    payload: dict[str, Any] = {"database_url": redact_database_url(database_url), "tables": {}}
    try:
        with engine.connect() as conn:
            for table_name in ["users", "sessions", "article_snapshots", "assignments", "publication_records"]:
                try:
                    rows = conn.execute(text(f"SELECT * FROM {table_name}")).mappings().all()
                except Exception as exc:  # noqa: BLE001
                    payload["tables"][table_name] = {"error": str(exc), "rows": []}
                    continue
                payload["tables"][table_name] = {"rows": [dict(row) for row in rows]}
    finally:
        engine.dispose()
    backup_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return {redact_database_url(database_url): str(backup_file)}


def redact_database_url(database_url: str) -> str:
    return re.sub(r"://([^:/@]+):([^@]+)@", r"://\1:***@", database_url)


def clean_existing_imported_items(project: dict[str, Any]) -> None:
    steps = project.setdefault("steps", {})
    article_state = steps.setdefault("article", {"status": "pending", "output": {}})
    output = article_state.setdefault("output", {})
    items = output.get("items") if isinstance(output.get("items"), list) else []
    output["items"] = [
        item for item in items
        if not (
            isinstance(item, dict)
            and isinstance(item.get("raw"), dict)
            and isinstance(item["raw"].get("imported_from"), dict)
            and item["raw"]["imported_from"].get("kind") == "robam_history"
        )
    ]


def write_projects(data_root: Path, articles: list[ArticleMeta]) -> dict[str, int]:
    now = utc_now()
    counts: dict[str, int] = {}
    by_project: dict[str, list[ArticleMeta]] = defaultdict(list)
    for article in articles:
        by_project[article.project_id].append(article)

    for project_id, project_articles in by_project.items():
        path = project_file(project_id, data_root)
        if not path.exists():
            raise FileNotFoundError(f"Project file not found: {path}")
        project = json.loads(path.read_text(encoding="utf-8"))
        clean_existing_imported_items(project)
        article_state = project.setdefault("steps", {}).setdefault("article", {"status": "pending", "output": {}})
        output = article_state.setdefault("output", {})
        items = output.get("items") if isinstance(output.get("items"), list) else []
        imported_items = []
        for article in project_articles:
            imported_items.append(
                {
                    "id": article.article_id,
                    "article_id": article.article_id,
                    "source_id": article.article_id,
                    "source_step": "imported",
                    "brief_id": "",
                    "keyword": article.keyword,
                    "type": article.article_type,
                    "article_type": article.article_type,
                    "title": article.title,
                    "role": "老板电器历史文章库导入",
                    "channel": "",
                    "status": "completed",
                    "used": "未使用",
                    "markdown": article.markdown,
                    "article_audit_status": "approved",
                    "article_audited_at": now,
                    "generated_at": now,
                    "updated_at": now,
                    "revision": 1,
                    "brief_revision": 1,
                    "raw": {
                        "imported_from": {
                            "kind": "robam_history",
                            "source_workbook": article.source_workbook,
                            "source_sheet": article.source_sheet,
                            "source_row": article.source_row,
                            "source_file": str(article.body_path) if article.body_path else "",
                            "source_file_value": article.source_file_value,
                            "match_method": article.match_method,
                            "match_score": article.match_score,
                            "placeholder": article.placeholder,
                            "content_hash": article.content_hash,
                        },
                        "article_library_row": article.raw,
                    },
                }
            )
        output["items"] = [item for item in items if isinstance(item, dict)] + imported_items
        output["status"] = "completed"
        output["updated_at"] = now
        article_state["status"] = "completed"
        article_state["error"] = None
        article_state["updated_at"] = now
        project["updated_at"] = now
        path.write_text(json.dumps(project, ensure_ascii=False, indent=2), encoding="utf-8")
        counts[project_id] = len(imported_items)
    return counts


def ensure_admin(conn: sqlite3.Connection, admin_username: str, admin_password: str, admin_display_name: str) -> str:
    row = conn.execute("SELECT id FROM users WHERE username = ?", (admin_username,)).fetchone()
    if row:
        return str(row["id"])
    admin_id = uuid.uuid4().hex
    now = utc_now()
    conn.execute(
        """
        INSERT INTO users (id, username, display_name, role, password_hash, active, created_at, updated_at)
        VALUES (?, ?, ?, 'admin', ?, 1, ?, ?)
        """,
        (admin_id, admin_username, admin_display_name, hash_password(admin_password), now, now),
    )
    return admin_id


def employee_id_for_name(name: str) -> str:
    return f"employee-{slug_hash(name, size=20)}"


def employee_username(name: str) -> str:
    return EMPLOYEE_USERNAME.get(name) or f"employee_{slug_hash(name, size=8)}"


def write_publishing_db(
    db_path: Path,
    articles: list[ArticleMeta],
    linked_publications: list[tuple[PublicationRow, ArticleMeta, str, float]],
    admin_username: str,
    admin_password: str,
    admin_display_name: str,
) -> dict[str, int]:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        now = utc_now()
        admin_id = ensure_admin(conn, admin_username, admin_password, admin_display_name)
        conn.execute("DELETE FROM sessions")
        conn.execute("DELETE FROM publication_records")
        conn.execute("DELETE FROM assignments")
        conn.execute("DELETE FROM article_snapshots")
        conn.execute("DELETE FROM users WHERE id != ?", (admin_id,))

        employee_names = sorted({publication.employee_name for publication, _, _, _ in linked_publications if publication.employee_name and publication.employee_name != "系统管理员"})
        employee_ids: dict[str, str] = {}
        for name in employee_names:
            employee_id = employee_id_for_name(name)
            employee_ids[name] = employee_id
            conn.execute(
                """
                INSERT INTO users (id, username, display_name, role, password_hash, active, created_at, updated_at)
                VALUES (?, ?, ?, 'employee', ?, 1, ?, ?)
                """,
                (employee_id, employee_username(name), name, hash_password("imported123"), now, now),
            )

        seen_article_ids: set[str] = set()
        for article in articles:
            if article.article_id in seen_article_ids:
                continue
            seen_article_ids.add(article.article_id)
            conn.execute(
                """
                INSERT INTO article_snapshots (
                    article_id, project_id, project_name, source_id, brief_id, keyword, article_type,
                    title, markdown, content_hash, article_audited_at, writing_updated_at, synced_at, active
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                """,
                (
                    article.article_id,
                    article.project_id,
                    article.project_name,
                    article.article_id,
                    "",
                    article.keyword,
                    article.article_type,
                    article.title,
                    article.markdown,
                    article.content_hash,
                    now,
                    now,
                    now,
                ),
            )

        assignment_pairs = {
            (employee_ids.get(publication.employee_name), article.project_id)
            for publication, article, _, _ in linked_publications
            if employee_ids.get(publication.employee_name)
        }
        for employee_id, project_id in sorted(pair for pair in assignment_pairs if pair[0]):
            conn.execute(
                """
                INSERT INTO assignments (id, user_id, project_id, keywords_json, article_types_json, created_at, updated_at)
                VALUES (?, ?, ?, '[]', '[]', ?, ?)
                """,
                (uuid.uuid4().hex, employee_id, project_id, now, now),
            )

        for publication, article, match_method, match_score in linked_publications:
            employee_id = employee_ids.get(publication.employee_name, admin_id)
            note_payload = {
                "source": "robam_history_import",
                "publication_row": publication.source_row,
                "article_match_method": match_method,
                "article_match_score": round(match_score, 4),
                "raw": publication.raw,
            }
            conn.execute(
                """
                INSERT INTO publication_records (
                    id, article_id, employee_id, channel_type, media_kind, media_category, media_name,
                    target_ai_platforms_json, reference_url, publish_url, published_at, order_id,
                    actual_cost, order_status, note, article_content_hash, created_at, updated_at
                )
                VALUES (?, ?, ?, '自媒体', '自媒体', '自媒体', ?, ?, '', ?, ?, '', 0, 'published', ?, ?, ?, ?)
                """,
                (
                    uuid.uuid4().hex,
                    article.article_id,
                    employee_id,
                    publication.media_name,
                    json.dumps(publication.ai_platforms, ensure_ascii=False),
                    publication.publish_url,
                    publication.published_at or now,
                    json.dumps(note_payload, ensure_ascii=False),
                    article.content_hash,
                    now,
                    now,
                ),
            )
        conn.commit()
        return {
            "users": 1 + len(employee_names),
            "employees": len(employee_names),
            "article_snapshots": len(seen_article_ids),
            "assignments": len(assignment_pairs),
            "publication_records": len(linked_publications),
        }
    finally:
        conn.close()


def write_publishing_sql(
    database_url: str,
    articles: list[ArticleMeta],
    linked_publications: list[tuple[PublicationRow, ArticleMeta, str, float]],
    admin_username: str,
    admin_password: str,
    admin_display_name: str,
) -> dict[str, int]:
    from sqlalchemy import create_engine, text

    engine = create_engine(database_url, future=True, pool_pre_ping=True)
    try:
        now = utc_now()
        with engine.begin() as conn:
            row = conn.execute(text("SELECT id FROM users WHERE username = :username"), {"username": admin_username}).mappings().first()
            if row:
                admin_id = str(row["id"])
            else:
                admin_id = uuid.uuid4().hex
                conn.execute(
                    text(
                        """
                        INSERT INTO users (id, username, display_name, role, password_hash, active, created_at, updated_at)
                        VALUES (:id, :username, :display_name, 'admin', :password_hash, 1, :created_at, :updated_at)
                        """
                    ),
                    {
                        "id": admin_id,
                        "username": admin_username,
                        "display_name": admin_display_name,
                        "password_hash": hash_password(admin_password),
                        "created_at": now,
                        "updated_at": now,
                    },
                )

            conn.execute(text("DELETE FROM sessions"))
            conn.execute(text("DELETE FROM publication_records"))
            conn.execute(text("DELETE FROM assignments"))
            conn.execute(text("DELETE FROM article_snapshots"))
            conn.execute(text("DELETE FROM users WHERE id != :admin_id"), {"admin_id": admin_id})

            employee_names = sorted({publication.employee_name for publication, _, _, _ in linked_publications if publication.employee_name and publication.employee_name != "系统管理员"})
            employee_ids: dict[str, str] = {}
            for name in employee_names:
                employee_id = employee_id_for_name(name)
                employee_ids[name] = employee_id
                conn.execute(
                    text(
                        """
                        INSERT INTO users (id, username, display_name, role, password_hash, active, created_at, updated_at)
                        VALUES (:id, :username, :display_name, 'employee', :password_hash, 1, :created_at, :updated_at)
                        """
                    ),
                    {
                        "id": employee_id,
                        "username": employee_username(name),
                        "display_name": name,
                        "password_hash": hash_password("imported123"),
                        "created_at": now,
                        "updated_at": now,
                    },
                )

            seen_article_ids: set[str] = set()
            article_rows = []
            for article in articles:
                if article.article_id in seen_article_ids:
                    continue
                seen_article_ids.add(article.article_id)
                article_rows.append(
                    {
                        "article_id": article.article_id,
                        "project_id": article.project_id,
                        "project_name": article.project_name,
                        "source_id": article.article_id,
                        "brief_id": "",
                        "keyword": article.keyword,
                        "article_type": article.article_type,
                        "title": article.title,
                        "markdown": article.markdown,
                        "content_hash": article.content_hash,
                        "article_audited_at": now,
                        "writing_updated_at": now,
                        "synced_at": now,
                    }
                )
            if article_rows:
                conn.execute(
                    text(
                        """
                        INSERT INTO article_snapshots (
                            article_id, project_id, project_name, source_id, brief_id, keyword, article_type,
                            title, markdown, content_hash, article_audited_at, writing_updated_at, synced_at, active
                        )
                        VALUES (
                            :article_id, :project_id, :project_name, :source_id, :brief_id, :keyword, :article_type,
                            :title, :markdown, :content_hash, :article_audited_at, :writing_updated_at, :synced_at, 1
                        )
                        """
                    ),
                    article_rows,
                )

            assignment_pairs = {
                (employee_ids.get(publication.employee_name), article.project_id)
                for publication, article, _, _ in linked_publications
                if employee_ids.get(publication.employee_name)
            }
            assignment_rows = [
                {
                    "id": uuid.uuid4().hex,
                    "user_id": employee_id,
                    "project_id": project_id,
                    "created_at": now,
                    "updated_at": now,
                }
                for employee_id, project_id in sorted(pair for pair in assignment_pairs if pair[0])
            ]
            if assignment_rows:
                conn.execute(
                    text(
                        """
                        INSERT INTO assignments (id, user_id, project_id, keywords_json, article_types_json, created_at, updated_at)
                        VALUES (:id, :user_id, :project_id, '[]', '[]', :created_at, :updated_at)
                        """
                    ),
                    assignment_rows,
                )

            record_rows = []
            for publication, article, match_method, match_score in linked_publications:
                note_payload = {
                    "source": "robam_history_import",
                    "publication_row": publication.source_row,
                    "article_match_method": match_method,
                    "article_match_score": round(match_score, 4),
                    "raw": publication.raw,
                }
                record_rows.append(
                    {
                        "id": uuid.uuid4().hex,
                        "article_id": article.article_id,
                        "employee_id": employee_ids.get(publication.employee_name, admin_id),
                        "media_name": publication.media_name,
                        "target_ai_platforms_json": json.dumps(publication.ai_platforms, ensure_ascii=False),
                        "publish_url": publication.publish_url,
                        "published_at": publication.published_at or now,
                        "note": json.dumps(note_payload, ensure_ascii=False),
                        "article_content_hash": article.content_hash,
                        "created_at": now,
                        "updated_at": now,
                    }
                )
            if record_rows:
                conn.execute(
                    text(
                        """
                        INSERT INTO publication_records (
                            id, article_id, employee_id, channel_type, media_kind, media_category, media_name,
                            target_ai_platforms_json, reference_url, publish_url, published_at, order_id,
                            actual_cost, order_status, note, article_content_hash, created_at, updated_at
                        )
                        VALUES (
                            :id, :article_id, :employee_id, '自媒体', '自媒体', '自媒体', :media_name,
                            :target_ai_platforms_json, '', :publish_url, :published_at, '',
                            0, 'published', :note, :article_content_hash, :created_at, :updated_at
                        )
                        """
                    ),
                    record_rows,
                )
        return {
            "users": 1 + len(employee_names),
            "employees": len(employee_names),
            "article_snapshots": len(seen_article_ids),
            "assignments": len(assignment_pairs),
            "publication_records": len(linked_publications),
        }
    finally:
        engine.dispose()


def load_active_articles_sql(database_url: str) -> list[ArticleMeta]:
    from sqlalchemy import create_engine, text

    engine = create_engine(database_url, future=True, pool_pre_ping=True)
    try:
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    """
                    SELECT article_id, project_id, project_name, keyword, article_type, title, markdown, content_hash
                    FROM article_snapshots
                    WHERE active = 1
                    """
                )
            ).mappings().all()
        articles: list[ArticleMeta] = []
        for row in rows:
            project_name = normalize_text(row["project_name"])
            category = "洗碗机" if "洗碗机" in project_name else "高端厨电" if "高端厨电" in project_name else "名气"
            articles.append(
                ArticleMeta(
                    article_id=normalize_text(row["article_id"]),
                    project_id=normalize_text(row["project_id"]),
                    project_name=project_name,
                    category=category,
                    keyword=normalize_text(row["keyword"]),
                    article_type=normalize_text(row["article_type"]),
                    title=normalize_text(row["title"]),
                    source_title=normalize_text(row["title"]),
                    source_file_value="",
                    source_workbook="mysql",
                    source_sheet="article_snapshots",
                    source_row=0,
                    markdown=normalize_text(row["markdown"]),
                    content_hash=normalize_text(row["content_hash"]),
                )
            )
        return articles
    finally:
        engine.dispose()


def ensure_sql_employee(conn: Any, name: str, now: str) -> str:
    from sqlalchemy import text

    if not name or name == "系统管理员":
        row = conn.execute(text("SELECT id FROM users WHERE username = 'admin'")).mappings().first()
        if row:
            return str(row["id"])
    row = conn.execute(text("SELECT id FROM users WHERE display_name = :name"), {"name": name}).mappings().first()
    if row:
        return str(row["id"])
    employee_id = employee_id_for_name(name)
    exists = conn.execute(text("SELECT id FROM users WHERE id = :id"), {"id": employee_id}).mappings().first()
    if exists:
        return str(exists["id"])
    conn.execute(
        text(
            """
            INSERT INTO users (id, username, display_name, role, password_hash, active, created_at, updated_at)
            VALUES (:id, :username, :display_name, 'employee', :password_hash, 1, :created_at, :updated_at)
            """
        ),
        {
            "id": employee_id,
            "username": employee_username(name),
            "display_name": name,
            "password_hash": hash_password("imported123"),
            "created_at": now,
            "updated_at": now,
        },
    )
    return employee_id


def ensure_sql_assignment(conn: Any, user_id: str, project_id: str, now: str) -> None:
    from sqlalchemy import text

    row = conn.execute(
        text("SELECT id FROM assignments WHERE user_id = :user_id AND project_id = :project_id"),
        {"user_id": user_id, "project_id": project_id},
    ).first()
    if row:
        return
    conn.execute(
        text(
            """
            INSERT INTO assignments (id, user_id, project_id, keywords_json, article_types_json, created_at, updated_at)
            VALUES (:id, :user_id, :project_id, '[]', '[]', :created_at, :updated_at)
            """
        ),
        {
            "id": uuid.uuid4().hex,
            "user_id": user_id,
            "project_id": project_id,
            "created_at": now,
            "updated_at": now,
        },
    )


def write_web_publications_sql(
    database_url: str,
    linked_publications: list[tuple[WebPublicationRow, ArticleMeta, str, float]],
) -> dict[str, Any]:
    from sqlalchemy import create_engine, text

    engine = create_engine(database_url, future=True, pool_pre_ping=True)
    try:
        now = utc_now()
        inserted = 0
        skipped_duplicate = 0
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM publication_records WHERE note LIKE '%robam_web_history_import%'"))
            for publication, article, match_method, match_score in linked_publications:
                duplicate = conn.execute(
                    text("SELECT id FROM publication_records WHERE article_id = :article_id AND publish_url = :publish_url"),
                    {"article_id": article.article_id, "publish_url": publication.publish_url},
                ).first()
                if duplicate:
                    skipped_duplicate += 1
                    continue
                employee_id = ensure_sql_employee(conn, publication.employee_name, now)
                ensure_sql_assignment(conn, employee_id, article.project_id, now)
                note_payload = {
                    "source": "robam_web_history_import",
                    "publication_row": publication.source_row,
                    "article_match_method": match_method,
                    "article_match_score": round(match_score, 4),
                    "publisher": publication.publisher,
                    "settled": publication.settled,
                    "raw": publication.raw,
                }
                conn.execute(
                    text(
                        """
                        INSERT INTO publication_records (
                            id, article_id, employee_id, channel_type, media_kind, media_category, media_name,
                            target_ai_platforms_json, reference_url, publish_url, published_at, order_id,
                            actual_cost, order_status, note, article_content_hash, created_at, updated_at
                        )
                        VALUES (
                            :id, :article_id, :employee_id, '网媒', :media_kind, :media_category, :media_name,
                            :target_ai_platforms_json, '', :publish_url, :published_at, '',
                            :actual_cost, 'published', :note, :article_content_hash, :created_at, :updated_at
                        )
                        """
                    ),
                    {
                        "id": uuid.uuid4().hex,
                        "article_id": article.article_id,
                        "employee_id": employee_id,
                        "media_kind": publication.media_kind,
                        "media_category": publication.media_category,
                        "media_name": publication.media_name,
                        "target_ai_platforms_json": json.dumps(publication.ai_platforms, ensure_ascii=False),
                        "publish_url": publication.publish_url,
                        "published_at": publication.published_at or now,
                        "actual_cost": publication.actual_cost,
                        "note": json.dumps(note_payload, ensure_ascii=False),
                        "article_content_hash": article.content_hash,
                        "created_at": now,
                        "updated_at": now,
                    },
                )
                inserted += 1
        return {"inserted": inserted, "skipped_duplicate": skipped_duplicate}
    finally:
        engine.dispose()


def summarize(
    articles: list[ArticleMeta],
    publications: list[PublicationRow],
    linked_publications: list[tuple[PublicationRow, ArticleMeta, str, float]],
) -> dict[str, Any]:
    placeholders = [article for article in articles if article.placeholder]
    low_confidence_publications = [
        {
            "publication_row": publication.source_row,
            "title": publication.titles[0] if publication.titles else "",
            "category": publication.category,
            "article_id": article.article_id,
            "method": method,
            "score": round(score, 4),
        }
        for publication, article, method, score in linked_publications
        if method not in {"title_exact", "title_exact_category"} or article.placeholder
    ]
    return {
        "articles": len(articles),
        "articles_by_project": dict(Counter(article.project_name for article in articles)),
        "articles_by_category": dict(Counter(article.category for article in articles)),
        "article_match_methods": dict(Counter(article.match_method for article in articles)),
        "placeholder_articles": len(placeholders),
        "placeholder_samples": [
            {
                "article_id": article.article_id,
                "title": article.title,
                "category": article.category,
                "keyword": article.keyword,
                "article_type": article.article_type,
                "source_workbook": article.source_workbook,
                "source_row": article.source_row,
                "best_score": round(article.match_score, 4),
            }
            for article in placeholders[:20]
        ],
        "publications": len(publications),
        "publication_links": len(linked_publications),
        "publication_match_methods": dict(Counter(method for _, _, method, _ in linked_publications)),
        "employees": sorted({publication.employee_name for publication in publications}),
        "low_confidence_publication_samples": low_confidence_publications[:30],
    }


def link_web_publications(
    publications: list[WebPublicationRow],
    articles: list[ArticleMeta],
    low_threshold: float,
) -> list[tuple[WebPublicationRow, ArticleMeta, str, float]]:
    article_index = index_articles(articles)
    linked: list[tuple[WebPublicationRow, ArticleMeta, str, float]] = []
    missing: list[str] = []
    for publication in publications:
        article, method, score = find_article_for_publication(publication, articles, article_index, low_threshold)  # type: ignore[arg-type]
        if not article:
            missing.append(f"row {publication.source_row}: {publication.titles[0] if publication.titles else publication.publish_url}")
            continue
        linked.append((publication, article, method, score))
    if missing:
        raise SystemExit("Web publication rows failed to match articles:\n" + "\n".join(missing[:50]))
    return linked


def summarize_web(
    publications: list[WebPublicationRow],
    linked_publications: list[tuple[WebPublicationRow, ArticleMeta, str, float]],
) -> dict[str, Any]:
    return {
        "web_publications": len(publications),
        "linked_web_publications": len(linked_publications),
        "total_actual_cost": sum(publication.actual_cost for publication in publications),
        "media_categories": dict(Counter(publication.media_category for publication in publications)),
        "raw_channel_categories": dict(Counter(publication.raw.get("发布渠道类别", "") for publication in publications)),
        "categories": dict(Counter(publication.category for publication in publications)),
        "creators": dict(Counter(publication.employee_name for publication in publications)),
        "match_methods": dict(Counter(method for _, _, method, _ in linked_publications)),
        "missing_links": sum(1 for publication in publications if not publication.publish_url),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Robam historical article libraries and publication records.")
    parser.add_argument("--attachments", type=Path, default=DEFAULT_ATTACHMENTS)
    parser.add_argument("--publication-stats", type=Path, default=DEFAULT_PUBLICATION_STATS)
    parser.add_argument("--web-publication-stats", type=Path, default=None)
    parser.add_argument("--library-xlsx", type=Path, action="append", default=[])
    parser.add_argument("--data-root", type=Path, default=None)
    parser.add_argument("--publishing-db", type=Path, default=None)
    parser.add_argument("--database-url", default="")
    parser.add_argument("--backup-root", type=Path, default=None)
    parser.add_argument("--low-threshold", type=float, default=0.60)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--rebuild", action="store_true", help="Required for writes; documents destructive rebuild intent.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    data_root = args.data_root or (ROOT / "app-data")
    publishing_db = args.publishing_db or (ROOT / "app-data" / "publishing" / "publishing.db")
    database_url = args.database_url or read_env_value(ROOT / ".env", "PUBLISHING_DATABASE_URL")
    backup_root = args.backup_root or (ROOT / "app-data" / "backups")
    library_paths = args.library_xlsx or DEFAULT_LIBRARY_XLSX

    if args.web_publication_stats:
        if not database_url:
            raise SystemExit("--web-publication-stats requires --database-url or PUBLISHING_DATABASE_URL in .env.")
        if not args.web_publication_stats.exists():
            raise SystemExit(f"Path not found: {args.web_publication_stats}")
        publications = read_web_publications(args.web_publication_stats)
        articles = load_active_articles_sql(database_url)
        linked_publications = link_web_publications(publications, articles, args.low_threshold)
        summary = summarize_web(publications, linked_publications)
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        if args.dry_run:
            print("Dry run complete. No database rows were modified.")
            return 0
        if not args.rebuild:
            raise SystemExit("Refusing to write without --rebuild.")
        backups = backup_sql_database(database_url, backup_root)
        result = copy.deepcopy(summary)
        result["backups"] = backups
        result["db_write_counts"] = write_web_publications_sql(database_url, linked_publications)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    for path in [args.attachments, args.publication_stats, *library_paths]:
        if not path.exists():
            raise SystemExit(f"Path not found: {path}")

    articles = build_articles(library_paths, args.attachments, args.low_threshold)
    publications = read_publications(args.publication_stats)
    articles, linked_publications = add_publication_placeholders(articles, publications, args.low_threshold)
    summary = summarize(articles, publications, linked_publications)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if args.dry_run:
        print("Dry run complete. No files or database rows were modified.")
        return 0
    if not args.rebuild:
        raise SystemExit("Refusing to write without --rebuild.")

    backups = backup_paths(data_root, publishing_db, backup_root)
    if database_url:
        backups.update(backup_sql_database(database_url, backup_root))
    project_counts = write_projects(data_root, articles)
    if database_url:
        db_counts = write_publishing_sql(
            database_url,
            articles,
            linked_publications,
            admin_username="admin",
            admin_password="admin123",
            admin_display_name="系统管理员",
        )
    else:
        db_counts = write_publishing_db(
            publishing_db,
            articles,
            linked_publications,
            admin_username="admin",
            admin_password="admin123",
            admin_display_name="系统管理员",
        )
    result = copy.deepcopy(summary)
    result["backups"] = backups
    result["project_write_counts"] = project_counts
    result["db_write_counts"] = db_counts
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
