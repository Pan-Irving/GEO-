import csv
import json
import shutil
import subprocess
import zipfile
from io import StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Callable
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader


SUPPORTED_EXTENSIONS = {
    ".md",
    ".txt",
    ".json",
    ".csv",
    ".xlsx",
    ".xlsm",
    ".xls",
    ".docx",
    ".docm",
    ".doc",
    ".rtf",
    ".odt",
    ".ods",
    ".html",
    ".htm",
    ".wordml",
    ".webarchive",
    ".pptx",
    ".pptm",
    ".ppsx",
    ".ppt",
    ".pps",
    ".pdf",
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
DOCX_EXTENSIONS = {".docx", ".docm"}
OFFICE_CONVERTER_EXTENSIONS = {
    ".doc",
    ".docx",
    ".docm",
    ".rtf",
    ".odt",
    ".ods",
    ".html",
    ".htm",
    ".wordml",
    ".webarchive",
    ".pptx",
    ".pptm",
    ".ppsx",
    ".ppt",
    ".pps",
    ".xls",
}


class ParseError(ValueError):
    pass


def parse_material(
    path: Path,
    *,
    image_ocr: Callable[[Path], str] | None = None,
    pdf_ocr: Callable[[Path], str] | None = None,
    pdf_page_ocr: Callable[[Path, list[int]], dict[int, str]] | None = None,
    pdf_ocr_max_pages: int | None = None,
) -> str:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ParseError(f"暂不支持该文件格式：{suffix}")
    if suffix in {".md", ".txt"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return f"```json\n{json.dumps(data, ensure_ascii=False, indent=2)}\n```"
    if suffix == ".csv":
        return parse_csv(path)
    if suffix in EXCEL_EXTENSIONS:
        try:
            return parse_xlsx(path)
        except Exception:
            if suffix in OFFICE_CONVERTER_EXTENSIONS:
                return parse_office_text(path)
            raise
    if suffix in DOCX_EXTENSIONS:
        try:
            return parse_docx(path)
        except ParseError:
            if suffix in OFFICE_CONVERTER_EXTENSIONS:
                return parse_office_text(path)
            raise
    if suffix in OFFICE_CONVERTER_EXTENSIONS:
        return parse_office_text(path)
    if suffix == ".pdf":
        return parse_pdf(path, pdf_ocr=pdf_ocr, pdf_page_ocr=pdf_page_ocr, pdf_ocr_max_pages=pdf_ocr_max_pages)
    if suffix in IMAGE_EXTENSIONS:
        return image_ocr(path) if image_ocr else parse_image_placeholder(path)
    raise ParseError(f"暂不支持该文件格式：{suffix}")


def parse_csv(path: Path) -> str:
    content = path.read_text(encoding="utf-8-sig", errors="ignore")
    reader = csv.reader(StringIO(content))
    rows = list(reader)
    return rows_to_markdown(rows, heading=path.name)


def parse_xlsx(path: Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sections: list[str] = []
    for sheet in workbook.worksheets:
        rows = [[cell if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
        sections.append(rows_to_markdown(rows, heading=sheet.title))
    return "\n\n".join(sections)


def parse_office_text(path: Path) -> str:
    text = try_textutil(path)
    if text is None:
        text = try_libreoffice_text(path)
    if text is None:
        raise ParseError("当前机器缺少可用的文档转换器。请安装 macOS textutil 或 LibreOffice 后重试。")
    cleaned = text.strip()
    if not cleaned:
        return f"## {path.name}\n\n（未抽取到可读文本。）"
    return f"## {path.name}\n\n{cleaned}"


def try_textutil(path: Path) -> str | None:
    if not shutil.which("textutil"):
        return None
    try:
        completed = subprocess.run(
            ["textutil", "-convert", "txt", "-stdout", str(path)],
            check=True,
            capture_output=True,
            text=True,
        )
    except (FileNotFoundError, subprocess.CalledProcessError):
        return None
    return completed.stdout


def try_libreoffice_text(path: Path) -> str | None:
    tool = shutil.which("soffice") or shutil.which("libreoffice")
    if not tool:
        return None
    try:
        with TemporaryDirectory() as temp_dir:
            subprocess.run(
                [tool, "--headless", "--convert-to", "txt:Text", "--outdir", temp_dir, str(path)],
                check=True,
                capture_output=True,
                text=True,
            )
            converted = Path(temp_dir) / f"{path.stem}.txt"
            if not converted.exists():
                candidates = sorted(Path(temp_dir).glob("*.txt"))
                if not candidates:
                    return None
                converted = candidates[0]
            return converted.read_text(encoding="utf-8", errors="ignore")
    except (FileNotFoundError, subprocess.CalledProcessError):
        return None


def parse_docx(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile) as exc:
        raise ParseError("DOCX 文件结构无效，无法读取 word/document.xml。") from exc

    root = ElementTree.fromstring(xml)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    blocks: list[str] = [f"## {path.name}", ""]
    for child in root.findall(".//w:body/*", namespace):
        tag = strip_xml_namespace(child.tag)
        if tag == "p":
            paragraph = docx_paragraph_text(child, namespace)
            if paragraph:
                blocks.append(paragraph)
        elif tag == "tbl":
            rows = docx_table_rows(child, namespace)
            if rows:
                blocks.append(rows_to_markdown(rows, heading="表格"))
    parsed = "\n\n".join(blocks).strip()
    if parsed == f"## {path.name}":
        return f"## {path.name}\n\n（DOCX 未抽取到可读文本。）"
    return parsed


def docx_paragraph_text(element: ElementTree.Element, namespace: dict[str, str]) -> str:
    texts = [node.text or "" for node in element.findall(".//w:t", namespace)]
    return "".join(texts).strip()


def docx_table_rows(element: ElementTree.Element, namespace: dict[str, str]) -> list[list[object]]:
    rows: list[list[object]] = []
    for row in element.findall(".//w:tr", namespace):
        cells = []
        for cell in row.findall("./w:tc", namespace):
            paragraphs = [docx_paragraph_text(paragraph, namespace) for paragraph in cell.findall(".//w:p", namespace)]
            cells.append("\n".join([paragraph for paragraph in paragraphs if paragraph]))
        if any(cells):
            rows.append(cells)
    return rows


def strip_xml_namespace(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def parse_pdf(
    path: Path,
    *,
    pdf_ocr: Callable[[Path], str] | None = None,
    pdf_page_ocr: Callable[[Path, list[int]], dict[int, str]] | None = None,
    pdf_ocr_max_pages: int | None = None,
) -> str:
    reader = PdfReader(str(path))
    sections: list[str] = [f"## {path.name}", ""]
    page_texts: list[tuple[int, str]] = []
    empty_page_indexes: list[int] = []
    for page_index, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        text = text.strip()
        page_texts.append((page_index, text))
        if not text:
            empty_page_indexes.append(page_index)

    ocr_results: dict[int, str] = {}
    ocr_page_indexes = empty_page_indexes
    if pdf_ocr_max_pages is not None:
        ocr_page_indexes = empty_page_indexes[: max(pdf_ocr_max_pages, 0)]
    if ocr_page_indexes and pdf_page_ocr:
        ocr_results = pdf_page_ocr(path, ocr_page_indexes)
    elif not any(text for _, text in page_texts) and pdf_ocr:
        return pdf_ocr(path)

    for page_index, text in page_texts:
        page_number = page_index + 1
        if text:
            sections.append(f"### 第 {page_number} 页\n\n{text}")
        elif ocr_results.get(page_index):
            sections.append(f"### 第 {page_number} 页 OCR\n\n{ocr_results[page_index]}")
    if len(sections) <= 2:
        sections.append("（PDF 未抽取到可读文本，可能是扫描件或图片型 PDF。请补充 OCR 后的文字资料。）")
    skipped_ocr_pages = len(empty_page_indexes) - len(ocr_page_indexes)
    if skipped_ocr_pages > 0 and pdf_page_ocr:
        sections.append(f"（为控制速度，剩余 {skipped_ocr_pages} 个无文本页面未 OCR。可切换完整 OCR 或调高 LOCAL_OCR_MAX_PAGES 后重新解析。）")
    return "\n\n".join(sections)


def parse_image_placeholder(path: Path) -> str:
    return (
        f"## {path.name}\n\n"
        "（图片资料已纳入资料池；本次未执行本地 OCR。"
        "如图片包含证书、截图、市场地位声明或产品参数，请补充对应文字说明，"
        "否则 Agent 只能把它作为存在该图片资料的线索，不能当作可引用证据。）"
    )


def rows_to_markdown(rows: list[list[object]], heading: str) -> str:
    normalized = [[str(cell).strip() for cell in row] for row in rows if any(str(cell).strip() for cell in row)]
    if not normalized:
        return f"## {heading}\n\n（空表）"
    width = max(len(row) for row in normalized)
    padded = [row + [""] * (width - len(row)) for row in normalized]
    header = padded[0]
    body = padded[1:]
    lines = [f"## {heading}", "", "| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]
    lines.extend("| " + " | ".join(row) + " |" for row in body)
    return "\n".join(lines)
